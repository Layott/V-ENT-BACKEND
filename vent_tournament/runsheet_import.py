"""Reading a run of show out of the document it was actually written in.

Nobody writes a run of show in a web form. They write it in a spreadsheet, they
rewrite it four times in the week before the show, and the last rewrite happens
in a car on the way to the venue. So the import is not a convenience: it is the
only way this feature survives contact with a real production.

Two doors, because both are how a spreadsheet leaves a spreadsheet:

**A file.** `.xlsx` straight out of Excel, Numbers or Google Sheets. Every
worksheet that looks like a running order becomes a day; the ones that do not,
and the CEO's own file has four of them (STREAM ELEMENTS, LEGEND, ROLES,
PLANS), are left alone rather than imported as gibberish.

**Pasted text.** CSV or TSV. Selecting a block of cells and pressing copy puts
tab separated text on the clipboard, which is the fastest path from a sheet
somebody is already looking at to a screen, and it needs no file at all.

## What it reads

The columns are the ones the CEO's sheet already uses, matched by name and not
by position, because a column gets inserted the week of the show and a position
based reader then silently shifts every value one to the left:

    PHASE | ACTIVITY | OWNS IT | MATCH | STARTS | ENDS | MINS

Names are matched loosely (`OWNS IT`, `OWNER`, `OWNS`, `WHO`), because the next
production will not use exactly these words.

## Two things it is careful about

**A blank PHASE continues the band above.** That is what a merged cell means in
a spreadsheet, and it is how a person reads the sheet. Carried forward at import
so the API answers with the band each row is genuinely in, rather than making
every reader reimplement the merge.

**Red bold means not confirmed.** The sheet's own LEGEND tab says so:
"RED BOLD TEXT means not confirmed. Scheduled and costed, not booked." That is
the single most useful fact on a run sheet, and in the source document it exists
only as a colour. Reading it off the font and storing it as a column is the
difference between carrying that meaning across and losing it at the door.
"""
import csv
import io
import re
from datetime import date, datetime, time, timedelta
from decimal import Decimal, InvalidOperation


# Every spelling of each column we are willing to accept, lowercased and with
# punctuation stripped. Matched loosely because the next sheet will not use
# exactly the CEO's words.
COLUMNS = {
    'phase': ('phase', 'segment', 'block', 'band'),
    'activity': ('activity', 'item', 'what', 'cue', 'segment name',
                 'description', 'action'),
    'owner': ('owns it', 'owns', 'owner', 'who', 'responsible', 'lead',
              'department', 'dept'),
    'match': ('match', 'fixture', 'game'),
    'starts_at': ('starts', 'start', 'start time', 'from', 'in', 'time'),
    'ends_at': ('ends', 'end', 'end time', 'to', 'out'),
    'minutes': ('mins', 'min', 'minutes', 'duration', 'length', 'runtime'),
    'note': ('note', 'notes', 'comment', 'comments', 'detail', 'details'),
}

# A row of headers has to carry at least this much to be a run of show. An
# activity column alone is enough: a sheet with nothing but a list of cues in
# order is still a run of show, and refusing it would refuse the simplest case.
REQUIRED = ('activity',)

MAX_ITEMS_PER_DAY = 600
MAX_DAYS = 30


def _clean(value):
    if value is None:
        return ''
    return re.sub(r'\s+', ' ', str(value)).strip()


def _header_key(value):
    """`OWNS IT` and `Owns it:` and `OWNS_IT` are the same column."""
    text = _clean(value).lower()
    text = re.sub(r'[^a-z0-9 ]+', ' ', text)
    return re.sub(r'\s+', ' ', text).strip()


def map_headers(cells):
    """Which column holds what, or None when this is not a header row.

    Returns `{field: index}`. A name we do not recognise is dropped rather than
    guessed at: a column called BAYSE is not a duration however hopefully it is
    read.
    """
    found = {}
    for index, cell in enumerate(cells):
        key = _header_key(cell)
        if not key:
            continue
        for field, names in COLUMNS.items():
            if field in found:
                continue
            if key in names:
                found[field] = index
                break
    if not all(field in found for field in REQUIRED):
        return None
    return found


def _as_time(value):
    """A time from whatever a spreadsheet decided that cell was.

    Excel stores a clock time as a fraction of a day, openpyxl usually hands
    back a `time`, a person pasting text hands back "10:00" or "10:00:00" or
    "10:00 AM", and any of the three arrives in the same column.
    """
    if value is None or value == '':
        return None
    if isinstance(value, time):
        return value.replace(second=0, microsecond=0)
    if isinstance(value, datetime):
        return value.time().replace(second=0, microsecond=0)
    if isinstance(value, timedelta):
        total = int(value.total_seconds()) % 86400
        return time(total // 3600, (total % 3600) // 60)
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        # A fraction of a day. Anything at or above 1 is a date serial with a
        # time in its fractional part, so take the fraction either way.
        fraction = float(value) % 1
        total = int(round(fraction * 86400))
        # Round to the minute; 10:00 stored as a float lands on 09:59:59.
        total = int(round(total / 60.0)) * 60
        total %= 86400
        return time(total // 3600, (total % 3600) // 60)

    text = _clean(value).upper().replace('.', ':')
    if not text:
        return None
    meridiem = None
    if text.endswith('AM') or text.endswith('PM'):
        meridiem = text[-2:]
        text = text[:-2].strip()
    match = re.match(r'^(\d{1,2})[:h]?(\d{2})?(?::(\d{2}))?$', text)
    if not match:
        return None
    hour = int(match.group(1))
    minute = int(match.group(2) or 0)
    if meridiem == 'PM' and hour < 12:
        hour += 12
    if meridiem == 'AM' and hour == 12:
        hour = 0
    if hour > 23 or minute > 59:
        return None
    return time(hour, minute)


def _as_minutes(value):
    if value is None or value == '':
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return Decimal(str(round(float(value), 1)))
    if isinstance(value, timedelta):
        return Decimal(str(round(value.total_seconds() / 60.0, 1)))
    text = _clean(value).lower().replace('mins', '').replace('min', '').strip()
    try:
        return Decimal(str(round(float(text), 1)))
    except (ValueError, TypeError, InvalidOperation):
        return None


def _minutes_between(starts, ends):
    if not starts or not ends:
        return None
    start = starts.hour * 60 + starts.minute
    end = ends.hour * 60 + ends.minute
    if end < start:
        # Past midnight. A show that ends at 00:30 ran thirty minutes past the
        # hour it started, not minus twenty three hours.
        end += 24 * 60
    return Decimal(str(float(end - start)))


# A date sitting inside a title row: "FRIDAY 2026-09-04", "5 September 2026",
# "04/09/2026". Read so a day imported from a sheet knows which day it is and
# the screen can then say what is on NOW.
_DATE_PATTERNS = (
    (r'(\d{4})-(\d{1,2})-(\d{1,2})', lambda m: date(int(m.group(1)), int(m.group(2)), int(m.group(3)))),
    (r'(\d{1,2})/(\d{1,2})/(\d{4})', lambda m: date(int(m.group(3)), int(m.group(2)), int(m.group(1)))),
)

_MONTHS = ('january', 'february', 'march', 'april', 'may', 'june', 'july',
           'august', 'september', 'october', 'november', 'december')


def find_date(text):
    """A date inside a free text line, or None.

    Deliberately conservative. A wrong date is worse than no date: it puts the
    NOW marker on the wrong day and tells everybody reading that the show is
    tomorrow.
    """
    if not text:
        return None
    lowered = _clean(text).lower()
    for pattern, build in _DATE_PATTERNS:
        match = re.search(pattern, lowered)
        if match:
            try:
                return build(match)
            except ValueError:
                return None
    # "4 September 2026" and "September 4, 2026".
    match = re.search(r'(\d{1,2})\s+([a-z]+)\s+(\d{4})', lowered)
    if match and match.group(2) in _MONTHS:
        try:
            return date(int(match.group(3)), _MONTHS.index(match.group(2)) + 1,
                        int(match.group(1)))
        except ValueError:
            return None
    match = re.search(r'([a-z]+)\s+(\d{1,2}),?\s+(\d{4})', lowered)
    if match and match.group(1) in _MONTHS:
        try:
            return date(int(match.group(3)), _MONTHS.index(match.group(1)) + 1,
                        int(match.group(2)))
        except ValueError:
            return None
    return None


# Red, in the several ways a spreadsheet says it. openpyxl hands back an ARGB
# string for a chosen colour and a theme index for one picked off the palette,
# and only the first can be read for hue without the workbook's theme, so this
# reads the first and treats everything else as confirmed. Erring towards
# confirmed is the safe direction: marking a booked segment unconfirmed on a
# public page is a claim about somebody else's business.
def _is_red(rgb):
    if not rgb or not isinstance(rgb, str) or len(rgb) < 6:
        return False
    body = rgb[-6:].upper()
    try:
        red = int(body[0:2], 16)
        green = int(body[2:4], 16)
        blue = int(body[4:6], 16)
    except ValueError:
        return False
    return red >= 0x90 and green <= 0x60 and blue <= 0x60


def _cell_unconfirmed(cell):
    """The LEGEND's rule, read off the font: red bold means not booked."""
    font = getattr(cell, 'font', None)
    if font is None:
        return False
    if not font.bold:
        return False
    colour = getattr(font, 'color', None)
    return _is_red(getattr(colour, 'rgb', None))


def _rows_to_items(rows, headers):
    """Cell rows into item dicts, carrying the phase band forward."""
    items = []
    phase = ''
    for cells, flags in rows:
        def at(field):
            index = headers.get(field)
            if index is None or index >= len(cells):
                return None
            return cells[index]

        activity = _clean(at('activity'))
        if not activity:
            continue

        this_phase = _clean(at('phase'))
        if this_phase:
            phase = this_phase

        starts = _as_time(at('starts_at'))
        ends = _as_time(at('ends_at'))
        minutes = _as_minutes(at('minutes'))
        if minutes is None:
            minutes = _minutes_between(starts, ends)

        items.append({
            'phase': phase[:80],
            'activity': activity[:400],
            'owner': _clean(at('owner'))[:120],
            'match': _clean(at('match'))[:120],
            'starts_at': starts,
            'ends_at': ends,
            'minutes': minutes,
            'note': _clean(at('note'))[:2000],
            'is_confirmed': not flags.get('unconfirmed', False),
            'position': len(items),
        })
        if len(items) >= MAX_ITEMS_PER_DAY:
            break
    return items


def from_worksheet(worksheet):
    """One worksheet into a day, or None when it is not a running order.

    Returns `{label, date, note, items}`.
    """
    header_row = None
    headers = None
    title_lines = []

    for index, row in enumerate(worksheet.iter_rows(min_row=1, max_row=12), start=1):
        values = [cell.value for cell in row]
        mapped = map_headers(values)
        if mapped:
            header_row = index
            headers = mapped
            break
        line = ' '.join(_clean(v) for v in values if _clean(v))
        if line:
            title_lines.append(line)

    if headers is None:
        return None

    rows = []
    for row in worksheet.iter_rows(min_row=header_row + 1):
        values = [cell.value for cell in row]
        if not any(_clean(v) for v in values):
            continue
        activity_index = headers.get('activity')
        flags = {}
        if activity_index is not None and activity_index < len(row):
            flags['unconfirmed'] = _cell_unconfirmed(row[activity_index])
        rows.append((values, flags))
        if len(rows) > MAX_ITEMS_PER_DAY * 2:
            break

    items = _rows_to_items(rows, headers)
    if not items:
        return None

    note = title_lines[0] if title_lines else ''
    return {
        'label': _clean(worksheet.title)[:80] or 'Day',
        'date': find_date(note),
        'note': note[:240],
        'items': items,
    }


def from_workbook(data):
    """Every worksheet in an xlsx that is a running order, in sheet order.

    Sheets that are not are skipped in silence rather than reported as an
    error: the CEO's own file carries four of them and they are all useful to
    the person who wrote it.
    """
    import openpyxl

    workbook = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    days = []
    for worksheet in workbook.worksheets:
        day = from_worksheet(worksheet)
        if day:
            days.append(day)
        if len(days) >= MAX_DAYS:
            break
    return days


def _sniff(text):
    """Tab or comma. Tab first, because a copy out of a spreadsheet is tabs and
    an activity like "MATCH LIVE, score bug, pop ups" is full of commas."""
    first = text.splitlines()[0] if text.splitlines() else ''
    return '\t' if first.count('\t') >= 1 else ','


def from_text(text, label='Day 1'):
    """Pasted CSV or TSV into one day.

    A header row is used when there is one. Without it the columns are read in
    the order the CEO's own sheet uses, which is the order somebody pasting a
    block out of that sheet will have.
    """
    text = (text or '').replace('\r\n', '\n').replace('\r', '\n').strip('\n')
    if not text.strip():
        return None

    reader = csv.reader(io.StringIO(text), delimiter=_sniff(text))
    rows = [row for row in reader if any(_clean(cell) for cell in row)]
    if not rows:
        return None

    headers = map_headers(rows[0])
    if headers:
        body = rows[1:]
    else:
        headers = {'phase': 0, 'activity': 1, 'owner': 2, 'match': 3,
                   'starts_at': 4, 'ends_at': 5, 'minutes': 6}
        body = rows

    items = _rows_to_items([(row, {}) for row in body], headers)
    if not items:
        return None
    return {'label': _clean(label)[:80] or 'Day 1', 'date': None,
            'note': '', 'items': items}
