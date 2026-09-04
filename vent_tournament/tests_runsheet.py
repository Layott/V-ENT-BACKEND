"""The run of show.

The fixture is the CEO's own file, `RIVALRY SERIES SEASON 2 EVENT FLOW.xlsx`,
Revision 2 of 4 September 2026, with the ROLES and PLANS tabs taken out because
they hold phone numbers and nothing the importer looks at. Everything else is
untouched, including the two sheets that are not running orders, because
skipping those correctly is half of what the import has to do.

Revision 2 arrived while this was being built and is the one that ships. It
imported with no change to the reader, which is the point of matching columns
by name rather than by position.

Testing against a fixture I wrote myself would prove the importer reads a sheet
shaped exactly the way I imagined a sheet is shaped. The whole risk here is that
a real spreadsheet is not.
"""
import os

from django.test import TestCase
from django.utils import timezone

from vent_auth.models import Users
from vent_event.models import Event

from .models import RunSheet, RunSheetDay, RunSheetItem, Tournament
from . import runsheet_import

FIXTURE = os.path.join(os.path.dirname(__file__), 'fixtures',
                       'rivalry-series-season-2-event-flow.xlsx')


def a_user(name):
    user = Users.objects.create(
        username=name, email='%s@vent.test' % name,
        login_session_token=('ros-%s' % name)[:20], is_active=True)
    user.login_session_created_at = timezone.now()
    user.login_session_2fa_at = timezone.now()
    user.save()
    return user, {'HTTP_AUTHORIZATION': 'Bearer %s' % user.login_session_token}


class Base(TestCase):
    def setUp(self):
        self.organiser, self.auth = a_user('org')
        self.stranger, self.stranger_auth = a_user('stranger')
        now = timezone.now()
        self.event = Event.objects.create(
            name='Rivalry Series Season 2', slug='rivalry-series-season-2',
            creator=self.organiser, event_type='physical', desc='x',
            entry_fee=0, start_date=now, end_date=now)

    def url(self, suffix=''):
        return '/event/%s/run-of-show/%s' % (self.event.slug, suffix)

    def upload(self, mode='replace'):
        with open(FIXTURE, 'rb') as handle:
            return self.client.post(
                self.url('import/'),
                data={'file': handle, 'mode': mode}, **self.auth)


# ---------------------------------------------------------------------------


class ImportTests(Base):
    """The CEO's actual spreadsheet, read end to end."""

    def test_the_real_workbook_imports_both_days(self):
        res = self.upload()
        self.assertEqual(res.status_code, 200, res.content)
        data = res.json()['data']
        self.assertEqual(data['imported_days'], 2)
        self.assertEqual(data['imported_items'], 161)

        days = data['sheet']['days']
        self.assertEqual([d['label'] for d in days], ['DAY 1', 'DAY 2'])
        self.assertEqual(days[0]['date'], '2026-09-04')
        self.assertEqual(days[1]['date'], '2026-09-05')

    def test_the_sheets_that_are_not_running_orders_are_skipped(self):
        """STREAM ELEMENTS and LEGEND are in the file and are not days.

        A file with four tabs of which two are a running order is the normal
        case, and importing the other two as gibberish would be worse than
        refusing the file.
        """
        self.upload()
        self.assertEqual(RunSheetDay.objects.count(), 2)
        self.assertFalse(
            RunSheetDay.objects.filter(label__icontains='LEGEND').exists())

    def test_the_first_cue_arrives_whole(self):
        self.upload()
        first = RunSheetItem.objects.order_by('day__position', 'position').first()
        self.assertEqual(first.activity, 'STARTING SOON, sponsor bed and ad reel')
        self.assertEqual(first.owner, 'GFX')
        self.assertEqual(first.phase, 'STREAM STARTS')
        self.assertEqual(first.starts_at.strftime('%H:%M'), '10:00')
        self.assertEqual(first.ends_at.strftime('%H:%M'), '10:10')
        self.assertEqual(float(first.minutes), 10.0)

    def test_a_blank_phase_continues_the_band_above(self):
        """Merged cells in a spreadsheet mean "same as above" to a reader.

        Row 2 of DAY 1 has an empty PHASE and belongs to STREAM STARTS. Leaving
        it blank would make every reader reimplement the merge, and they would
        each do it differently.
        """
        self.upload()
        day = RunSheetDay.objects.order_by('position').first()
        second = day.items.order_by('position')[1]
        self.assertEqual(second.activity, 'NATIONAL ANTHEMS, all five nations')
        self.assertEqual(second.phase, 'STREAM STARTS')

    def test_a_cue_carries_its_match(self):
        self.upload()
        row = RunSheetItem.objects.filter(match='NGA1 v GHA1').first()
        self.assertIsNotNone(row)

    def test_the_day_header_becomes_the_subtitle(self):
        self.upload()
        sheet = RunSheet.objects.get()
        self.assertIn('doors 10:00 to 18:00', sheet.subtitle)


class PasteTests(Base):
    """Copying a block of cells is how a sheet leaves a sheet fastest."""

    TSV = (
        'PHASE\tACTIVITY\tOWNS IT\tMATCH\tSTARTS\tENDS\tMINS\n'
        'STREAM STARTS\tSTARTING SOON\tGFX\t\t10:00\t10:10\t10\n'
        '\tNATIONAL ANTHEMS\tFloor / audio\t\t10:10\t10:20\t10\n'
        'MATCHES ONGOING\tMATCH LIVE\tCasters\tNGA1 v GHA1\t10:52\t11:17\t25\n'
    )

    def test_pasted_tab_separated_rows_import(self):
        res = self.client.post(self.url('import/'),
                               data={'text': self.TSV, 'label': 'Day 1'},
                               content_type='application/json', **self.auth)
        self.assertEqual(res.status_code, 200, res.content)
        self.assertEqual(res.json()['data']['imported_items'], 3)
        rows = list(RunSheetItem.objects.order_by('position'))
        self.assertEqual(rows[1].phase, 'STREAM STARTS')
        self.assertEqual(rows[2].match, 'NGA1 v GHA1')
        self.assertEqual(float(rows[2].minutes), 25.0)

    def test_commas_inside_an_activity_survive_a_tab_paste(self):
        """"MATCH LIVE, score bug, pop ups" is one cell, not three.

        Sniffing comma first would split it, which is why tab wins whenever a
        tab is present.
        """
        text = ('PHASE\tACTIVITY\tOWNS IT\tMATCH\tSTARTS\tENDS\tMINS\n'
                '\tMATCH LIVE, score bug, pop ups\tCasters\t\t10:52\t11:17\t25\n')
        self.client.post(self.url('import/'),
                         data={'text': text}, content_type='application/json',
                         **self.auth)
        row = RunSheetItem.objects.get()
        self.assertEqual(row.activity, 'MATCH LIVE, score bug, pop ups')

    def test_a_paste_with_no_header_reads_the_columns_in_order(self):
        text = 'STREAM STARTS\tSTARTING SOON\tGFX\t\t10:00\t10:10\t10\n'
        res = self.client.post(self.url('import/'), data={'text': text},
                               content_type='application/json', **self.auth)
        self.assertEqual(res.status_code, 200, res.content)
        row = RunSheetItem.objects.get()
        self.assertEqual(row.activity, 'STARTING SOON')
        self.assertEqual(row.owner, 'GFX')

    def test_nothing_that_looks_like_a_running_order_is_refused_with_a_reason(self):
        res = self.client.post(self.url('import/'), data={'text': '\n\n'},
                               content_type='application/json', **self.auth)
        self.assertEqual(res.status_code, 400)
        self.assertEqual(res.json()['code'], 'NOTHING_TO_IMPORT')

    def test_a_duration_is_worked_out_when_the_sheet_gives_only_times(self):
        text = ('ACTIVITY\tSTARTS\tENDS\n'
                'ANALYST DESK\t13:01\t13:06\n')
        self.client.post(self.url('import/'), data={'text': text},
                         content_type='application/json', **self.auth)
        self.assertEqual(float(RunSheetItem.objects.get().minutes), 5.0)


class ReimportTests(Base):
    """A sheet is rewritten four times in the week before a show."""

    def test_importing_twice_does_not_double_the_sheet(self):
        self.upload()
        self.upload()
        self.assertEqual(RunSheetDay.objects.count(), 2)
        self.assertEqual(RunSheetItem.objects.count(), 161)

    def test_append_adds_days_beside_what_is_there(self):
        self.upload()
        self.upload(mode='append')
        self.assertEqual(RunSheetDay.objects.count(), 4)

    def test_an_unknown_mode_is_refused(self):
        with open(FIXTURE, 'rb') as handle:
            res = self.client.post(self.url('import/'),
                                   data={'file': handle, 'mode': 'merge'},
                                   **self.auth)
        self.assertEqual(res.status_code, 400)
        self.assertEqual(res.json()['code'], 'INVALID_MODE')


class VisibilityTests(Base):
    def test_a_new_sheet_is_private(self):
        """The default cannot be anything else.

        A run sheet carries staff names, when the money is counted and which
        segments are not booked. Publishing it by accident is worse than not
        having the feature.
        """
        self.upload()
        self.assertEqual(RunSheet.objects.get().visibility, RunSheet.PRIVATE)

    def test_the_organiser_can_publish_it(self):
        self.upload()
        res = self.client.post(self.url(), data={'visibility': 'public'},
                               content_type='application/json', **self.auth)
        self.assertEqual(res.status_code, 200, res.content)
        self.assertEqual(RunSheet.objects.get().visibility, 'public')

    def test_a_visibility_that_is_not_one_of_the_three_is_refused(self):
        self.upload()
        res = self.client.post(self.url(), data={'visibility': 'everyone'},
                               content_type='application/json', **self.auth)
        self.assertEqual(res.status_code, 400)
        self.assertEqual(res.json()['code'], 'INVALID_VISIBILITY')

    def test_owners_are_withheld_at_the_api_when_the_organiser_hides_them(self):
        """Not hidden by CSS. A column dropped in the page is a column anybody
        can read."""
        self.upload()
        self.client.post(self.url(),
                         data={'visibility': 'public', 'show_owners': False},
                         content_type='application/json', **self.auth)
        res = self.client.get(self.url())
        item = res.json()['data']['sheet']['days'][0]['items'][0]
        self.assertNotIn('owner', item)

        # The organiser still sees them.
        mine = self.client.get(self.url(), **self.auth)
        self.assertEqual(
            mine.json()['data']['sheet']['days'][0]['items'][0]['owner'], 'GFX')


class AccessTests(Base):
    def token(self):
        return RunSheet.objects.get().token

    def test_a_private_sheet_is_404_to_a_stranger(self):
        """404 and not 403. A refusal confirms there is a run of show for this
        event, which is itself the thing being kept private."""
        self.upload()
        res = self.client.get(self.url())
        self.assertEqual(res.status_code, 404)
        signed_in = self.client.get(self.url(), **self.stranger_auth)
        self.assertEqual(signed_in.status_code, 404)

    def test_a_public_sheet_reads_signed_out(self):
        self.upload()
        self.client.post(self.url(), data={'visibility': 'public'},
                         content_type='application/json', **self.auth)
        res = self.client.get(self.url())
        self.assertEqual(res.status_code, 200, res.content)
        self.assertEqual(len(res.json()['data']['sheet']['days']), 2)
        self.assertFalse(res.json()['data']['can_manage'])

    def test_a_link_only_sheet_opens_by_token_and_not_by_address(self):
        self.upload()
        self.client.post(self.url(), data={'visibility': 'link'},
                         content_type='application/json', **self.auth)
        self.assertEqual(self.client.get(self.url()).status_code, 404)
        res = self.client.get('/run-of-show/%s/' % self.token())
        self.assertEqual(res.status_code, 200, res.content)

    def test_a_private_sheet_is_404_even_with_its_token(self):
        """Making a sheet private again has to close the address that was sent
        out, otherwise private means nothing once a link has been shared."""
        self.upload()
        token = self.token()
        self.client.post(self.url(), data={'visibility': 'link'},
                         content_type='application/json', **self.auth)
        self.client.post(self.url(), data={'visibility': 'private'},
                         content_type='application/json', **self.auth)
        self.assertEqual(
            self.client.get('/run-of-show/%s/' % token).status_code, 404)

    def test_a_token_nobody_issued_is_404(self):
        self.assertEqual(
            self.client.get('/run-of-show/not-a-real-token/').status_code, 404)

    def test_a_reader_is_never_told_the_token(self):
        """A public sheet handed out its token would hand out an address that
        keeps working after the organiser makes it private."""
        self.upload()
        self.client.post(self.url(), data={'visibility': 'public'},
                         content_type='application/json', **self.auth)
        res = self.client.get(self.url())
        self.assertIsNone(res.json()['data']['sheet']['token'])
        mine = self.client.get(self.url(), **self.auth)
        self.assertTrue(mine.json()['data']['sheet']['token'])

    def test_an_event_with_no_run_of_show_says_so_to_its_organiser(self):
        res = self.client.get(self.url(), **self.auth)
        self.assertEqual(res.status_code, 200)
        self.assertIsNone(res.json()['data']['sheet'])
        self.assertTrue(res.json()['data']['can_manage'])

    def test_an_event_with_no_run_of_show_is_404_to_everybody_else(self):
        self.assertEqual(self.client.get(self.url()).status_code, 404)


class WriteAccessTests(Base):
    """The API is what stops anybody. A hidden button is not a permission."""

    def setUp(self):
        super().setUp()
        self.upload()
        self.sheet = RunSheet.objects.get()
        self.day = self.sheet.days.first()
        self.item = self.day.items.first()

    def each_write(self):
        return [
            ('post', self.url(), {'visibility': 'public'}),
            ('post', self.url('import/'), {'text': 'ACTIVITY\nx\n'}),
            ('post', self.url('days/'), {'label': 'Day 3'}),
            ('patch', self.url('days/%s/' % self.day.id), {'label': 'X'}),
            ('delete', self.url('days/%s/' % self.day.id), None),
            ('post', self.url('items/'), {'activity': 'x'}),
            ('patch', self.url('items/%s/' % self.item.id), {'activity': 'x'}),
            ('delete', self.url('items/%s/' % self.item.id), None),
            ('delete', self.url(), None),
        ]

    def test_anonymous_is_refused_everywhere(self):
        for method, url, body in self.each_write():
            call = getattr(self.client, method)
            res = (call(url, data=body, content_type='application/json')
                   if body is not None else call(url))
            self.assertIn(res.status_code, (401, 403),
                          '%s %s answered %s' % (method, url, res.status_code))

    def test_a_signed_in_stranger_is_refused_everywhere(self):
        for method, url, body in self.each_write():
            call = getattr(self.client, method)
            res = (call(url, data=body, content_type='application/json',
                        **self.stranger_auth)
                   if body is not None
                   else call(url, **self.stranger_auth))
            self.assertEqual(res.status_code, 403,
                             '%s %s answered %s' % (method, url,
                                                    res.status_code))


class EditTests(Base):
    def setUp(self):
        super().setUp()
        self.upload()
        self.sheet = RunSheet.objects.get()
        self.day = self.sheet.days.first()

    def test_a_cue_is_added_corrected_and_removed(self):
        res = self.client.post(
            self.url('items/'),
            data={'day_id': self.day.id, 'activity': 'SPONSOR READ',
                  'owner': 'Host', 'starts_at': '12:05', 'ends_at': '12:07'},
            content_type='application/json', **self.auth)
        self.assertEqual(res.status_code, 200, res.content)
        row = RunSheetItem.objects.filter(activity='SPONSOR READ').get()
        self.assertEqual(float(row.minutes), 2.0)

        self.client.patch(self.url('items/%s/' % row.id),
                          data={'activity': 'SPONSOR READ, Bayse'},
                          content_type='application/json', **self.auth)
        row.refresh_from_db()
        self.assertEqual(row.activity, 'SPONSOR READ, Bayse')

        self.client.delete(self.url('items/%s/' % row.id), **self.auth)
        self.assertFalse(RunSheetItem.objects.filter(pk=row.id).exists())

    def test_a_cue_with_no_activity_is_refused(self):
        res = self.client.post(self.url('items/'), data={'activity': '  '},
                               content_type='application/json', **self.auth)
        self.assertEqual(res.status_code, 400)
        self.assertEqual(res.json()['code'], 'VALIDATION_FAILED')

    def test_a_time_that_cannot_be_read_says_how_to_write_it(self):
        res = self.client.post(
            self.url('items/'),
            data={'activity': 'x', 'starts_at': 'half past ten'},
            content_type='application/json', **self.auth)
        self.assertEqual(res.status_code, 400)
        self.assertEqual(res.json()['code'], 'INVALID_TIME')

    def test_a_day_is_added_renamed_and_removed(self):
        self.client.post(self.url('days/'),
                         data={'label': 'Day 3', 'date': '2026-09-06'},
                         content_type='application/json', **self.auth)
        day = RunSheetDay.objects.get(label='Day 3')
        self.assertEqual(day.date.isoformat(), '2026-09-06')

        self.client.patch(self.url('days/%s/' % day.id),
                          data={'label': 'Finals day'},
                          content_type='application/json', **self.auth)
        day.refresh_from_db()
        self.assertEqual(day.label, 'Finals day')

        self.client.delete(self.url('days/%s/' % day.id), **self.auth)
        self.assertFalse(RunSheetDay.objects.filter(pk=day.id).exists())

    def test_removing_the_sheet_removes_its_days_and_cues(self):
        self.client.delete(self.url(), **self.auth)
        self.assertEqual(RunSheet.objects.count(), 0)
        self.assertEqual(RunSheetDay.objects.count(), 0)
        self.assertEqual(RunSheetItem.objects.count(), 0)


class TournamentParityTests(TestCase):
    """The same six routes on the other thing V-ENT runs.

    Built at the same time and not left for later. Five times in one day on
    3 September a control turned out to exist on one of these two surfaces and
    not the other, which is the fault this whole class is here to make
    impossible.
    """

    def setUp(self):
        self.organiser, self.auth = a_user('t_org')
        now = timezone.now()
        self.tournament = Tournament.objects.create(
            tournament_title='Rivalry Series S2', slug='rivalry-series-s2',
            tournament_creator=self.organiser, tournament_type='online',
            start_date_and_time=now, end_date_and_time=now)

    def url(self, suffix=''):
        return '/tournament/%s/run-of-show/%s' % (self.tournament.slug, suffix)

    def test_a_tournament_imports_shares_and_reads_the_same_way(self):
        with open(FIXTURE, 'rb') as handle:
            res = self.client.post(self.url('import/'),
                                   data={'file': handle}, **self.auth)
        self.assertEqual(res.status_code, 200, res.content)
        self.assertEqual(res.json()['data']['imported_items'], 161)

        sheet = RunSheet.objects.get()
        self.assertIsNone(sheet.event_id)
        self.assertEqual(sheet.tournament_id, self.tournament.tournament_id)

        self.client.post(self.url(), data={'visibility': 'link'},
                         content_type='application/json', **self.auth)
        res = self.client.get('/run-of-show/%s/' % sheet.token)
        self.assertEqual(res.status_code, 200, res.content)
        self.assertEqual(res.json()['data']['sheet']['owner']['kind'],
                         'tournament')


class ParsingTests(TestCase):
    """The reader on its own, without a request in the way."""

    def test_a_time_is_read_however_the_sheet_stored_it(self):
        from datetime import time
        self.assertEqual(runsheet_import._as_time('10:00'), time(10, 0))
        self.assertEqual(runsheet_import._as_time('10:00:00'), time(10, 0))
        self.assertEqual(runsheet_import._as_time('1:05 PM'), time(13, 5))
        self.assertEqual(runsheet_import._as_time('12:30 AM'), time(0, 30))
        # Excel's fraction of a day. 10:00 is 0.41666..., which rounds to
        # 09:59:59 without the rounding to the minute.
        self.assertEqual(runsheet_import._as_time(10 / 24.0), time(10, 0))
        self.assertIsNone(runsheet_import._as_time('half ten'))
        self.assertIsNone(runsheet_import._as_time('25:00'))

    def test_a_duration_past_midnight_is_not_negative(self):
        from datetime import time
        minutes = runsheet_import._minutes_between(time(23, 50), time(0, 20))
        self.assertEqual(float(minutes), 30.0)

    def test_a_date_is_found_in_a_title_line(self):
        from datetime import date
        self.assertEqual(
            runsheet_import.find_date(
                'OVERALL EVENT FLOW, FRIDAY 2026-09-04, doors 10:00 to 18:00'),
            date(2026, 9, 4))
        self.assertEqual(runsheet_import.find_date('5 September 2026'),
                         date(2026, 9, 5))
        self.assertIsNone(runsheet_import.find_date('OVERALL EVENT FLOW'))

    def test_a_column_is_found_by_name_and_not_by_position(self):
        headers = runsheet_import.map_headers(
            ['MINS', 'ACTIVITY', 'OWNS IT', 'STARTS'])
        self.assertEqual(headers['activity'], 1)
        self.assertEqual(headers['minutes'], 0)
        self.assertEqual(headers['owner'], 2)

    def test_a_row_of_data_is_not_mistaken_for_a_header(self):
        self.assertIsNone(runsheet_import.map_headers(
            ['STREAM STARTS', 'STARTING SOON', 'GFX']))

    def test_red_bold_is_read_as_not_confirmed(self):
        """The sheet's LEGEND: "RED BOLD TEXT means not confirmed."

        In the source document that meaning exists only as a colour, so losing
        it at the door loses the most useful fact on the sheet.
        """
        import io
        import openpyxl
        from openpyxl.styles import Font

        book = openpyxl.Workbook()
        page = book.active
        page.title = 'DAY 1'
        page.append(['PHASE', 'ACTIVITY', 'OWNS IT'])
        page.append(['BREAK', 'CELEBRITY MATCH', 'Host'])
        page.append(['BREAK', 'OPENING SEQUENCE', 'Host'])
        page.cell(row=3, column=2).font = Font(bold=True, color='FFFF0000')
        buffer = io.BytesIO()
        book.save(buffer)

        days = runsheet_import.from_workbook(buffer.getvalue())
        items = days[0]['items']
        self.assertTrue(items[0]['is_confirmed'])
        self.assertFalse(items[1]['is_confirmed'])

    def test_bold_black_is_still_confirmed(self):
        """Erring towards confirmed is the safe direction: marking a booked
        segment unconfirmed on a public page is a claim about somebody else's
        business."""
        import io
        import openpyxl
        from openpyxl.styles import Font

        book = openpyxl.Workbook()
        page = book.active
        page.append(['ACTIVITY'])
        page.append(['DAY WRAP'])
        page.cell(row=2, column=1).font = Font(bold=True, color='FF000000')
        buffer = io.BytesIO()
        book.save(buffer)
        self.assertTrue(
            runsheet_import.from_workbook(buffer.getvalue())[0]['items'][0]
            ['is_confirmed'])
